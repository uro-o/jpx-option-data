import json
import re
from pathlib import Path
from datetime import datetime

import openpyxl


# ============================================================
# 設定
# ============================================================

DATA_DIR = Path("data")
OUTPUT_DIR = Path("dashboard/public")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# 最新Excelを取得
# ============================================================

files = sorted(
    DATA_DIR.glob("*_open_interest.xlsx")
)

if not files:
    raise FileNotFoundError(
        "dataフォルダにopen_interest.xlsxがありません。"
    )

latest_file = files[-1]

print("使用するExcel:")
print(latest_file)


# ============================================================
# Excel読み込み
# ============================================================

wb = openpyxl.load_workbook(
    latest_file,
    data_only=True
)

ws = wb["別紙1"]


# ============================================================
# 日付
# ============================================================

raw_date = ws["A2"].value

if isinstance(raw_date, datetime):
    data_date = raw_date.strftime("%Y-%m-%d")
else:
    data_date = str(raw_date)


# ============================================================
# オプション名解析
#
# 例:
# NIKKEI 225 P2609-70000
# NIKKEI 225 C2609-75000
# ============================================================

pattern = re.compile(
    r"NIKKEI 225 ([PC])(\d{4})-(\d+)"
)


def parse_option_name(name):

    if not name:
        return None

    match = pattern.search(str(name))

    if not match:
        return None

    option_type = match.group(1)
    expiry_code = match.group(2)
    strike = int(match.group(3))

    year = 2000 + int(expiry_code[:2])
    month = int(expiry_code[2:])

    expiry = f"{year}-{month:02d}"

    return {
        "type": option_type,
        "expiry": expiry,
        "strike": strike,
    }


# ============================================================
# データ格納
# ============================================================

options = []


# ============================================================
# 別紙1を解析
# ============================================================

for row in range(7, ws.max_row + 1):

    # -----------------------------
    # PUT
    # -----------------------------

    put_name = ws.cell(row, 1).value
    put_volume = ws.cell(row, 2).value
    put_oi = ws.cell(row, 3).value
    put_change = ws.cell(row, 4).value
    put_prev_oi = ws.cell(row, 5).value

    parsed = parse_option_name(put_name)

    if parsed:

        options.append({
            "type": "PUT",
            "expiry": parsed["expiry"],
            "strike": parsed["strike"],
            "volume": put_volume or 0,
            "oi": put_oi or 0,
            "change": put_change or 0,
            "prev_oi": put_prev_oi or 0,
        })


    # -----------------------------
    # CALL
    # -----------------------------

    call_name = ws.cell(row, 7).value
    call_volume = ws.cell(row, 8).value
    call_oi = ws.cell(row, 9).value
    call_change = ws.cell(row, 10).value
    call_prev_oi = ws.cell(row, 11).value

    parsed = parse_option_name(call_name)

    if parsed:

        options.append({
            "type": "CALL",
            "expiry": parsed["expiry"],
            "strike": parsed["strike"],
            "volume": call_volume or 0,
            "oi": call_oi or 0,
            "change": call_change or 0,
            "prev_oi": call_prev_oi or 0,
        })


# ============================================================
# Max Pain計算
# ============================================================

def calculate_max_pain(expiry):

    data = [
        x for x in options
        if x["expiry"] == expiry
    ]

    if not data:
        return None

    strikes = sorted(
        set(x["strike"] for x in data)
    )

    min_pain = None
    max_pain_strike = None

    for settlement in strikes:

        pain = 0

        for x in data:

            strike = x["strike"]
            oi = x["oi"]

            if x["type"] == "CALL":

                pain += max(
                    0,
                    settlement - strike
                ) * oi

            else:

                pain += max(
                    0,
                    strike - settlement
                ) * oi

        if min_pain is None or pain < min_pain:

            min_pain = pain
            max_pain_strike = settlement

    return max_pain_strike


# ============================================================
# 限月一覧
# ============================================================

expiries = sorted(
    set(x["expiry"] for x in options)
)


# ============================================================
# Max Pain
# ============================================================

max_pain = {}

for expiry in expiries:

    max_pain[expiry] = calculate_max_pain(
        expiry
    )


# ============================================================
# Strike一覧
# ============================================================

strikes_by_expiry = {}

for expiry in expiries:

    strikes = sorted(
        set(
            x["strike"]
            for x in options
            if x["expiry"] == expiry
        )
    )

    strikes_by_expiry[expiry] = strikes


# ============================================================
# JSON生成
# ============================================================

output = {

    "updated": data_date,

    "source_file": latest_file.name,

    "expiries": expiries,

    "max_pain": max_pain,

    "strikes": strikes_by_expiry,

    "options": options,

}


# ============================================================
# 保存
# ============================================================

output_file = OUTPUT_DIR / "data.json"

with open(
    output_file,
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        output,
        f,
        ensure_ascii=False,
        indent=2
    )


print("")
print("========================================")
print("Dashboard data generated")
print("========================================")

print("更新日:", data_date)
print("限月数:", len(expiries))
print("オプション数:", len(options))
print("Max Pain:")

for expiry, value in max_pain.items():

    print(
        " ",
        expiry,
        ":",
        value
    )

print("")
print("保存先:", output_file)
